from datetime import date, datetime, timedelta, timezone
from decimal import Decimal
from hmac import compare_digest
import os
from pathlib import Path

from fastapi import Depends, FastAPI, HTTPException, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import Workbook
from sqlalchemy import and_, extract, func
from sqlalchemy.orm import Session, joinedload

from database import Base, engine, get_db
from models import Ingredient, Purchase, PurchaseItem
from schemas import (
    DashboardOut,
    IngredientCreate,
    IngredientOut,
    IngredientUpdate,
    PurchaseCreate,
    PurchaseItemOut,
    PurchaseOut,
    ReportOut,
    ReportPoint,
    TopIngredient,
)


app = FastAPI(title="Restaurant Inventory API", version="1.0.0")

allowed_origins = [
    origin.strip()
    for origin in os.getenv(
        "FRONTEND_ORIGINS",
        "http://localhost:5173,http://127.0.0.1:5173",
    ).split(",")
    if origin.strip()
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

API_KEY = os.getenv("API_KEY", "")
APP_PASSWORD = os.getenv("APP_PASSWORD", "")
AUTH_DISABLED = os.getenv("DISABLE_AUTH", "").lower() == "true"
PUBLIC_PATHS = {"/health"}


@app.middleware("http")
async def require_app_auth(request: Request, call_next):
    if request.method == "OPTIONS" or request.url.path in PUBLIC_PATHS:
        return await call_next(request)
    if AUTH_DISABLED:
        return await call_next(request)
    if not API_KEY or not APP_PASSWORD:
        return JSONResponse({"detail": "API auth is not configured"}, status_code=503)
    if API_KEY and not compare_digest(request.headers.get("x-api-key", ""), API_KEY):
        return JSONResponse({"detail": "Invalid API key"}, status_code=401)
    if APP_PASSWORD and not compare_digest(request.headers.get("x-app-password", ""), APP_PASSWORD):
        return JSONResponse({"detail": "Invalid app password"}, status_code=401)
    return await call_next(request)

DEFAULT_INGREDIENTS = [
    ("หมู", "กิโลกรัม"),
    ("ไก่", "กิโลกรัม"),
    ("เนื้อ", "กิโลกรัม"),
    ("กุ้ง", "กิโลกรัม"),
    ("ปลาหมึก", "กิโลกรัม"),
    ("ไข่", "ฟอง"),
    ("ข้าวสาร", "ถุง"),
    ("พริก", "กรัม"),
    ("กระเทียม", "กรัม"),
    ("น้ำปลา", "ขวด"),
    ("น้ำมันพืช", "ขวด"),
]


def decimal_value(value):
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def money(value):
    return decimal_value(value).quantize(Decimal("0.01"))


def serialize_purchase(purchase: Purchase) -> PurchaseOut:
    return PurchaseOut(
        id=purchase.id,
        purchase_date=purchase.purchase_date,
        total_amount=money(purchase.total_amount),
        created_at=purchase.created_at,
        updated_at=purchase.updated_at,
        items=[
            PurchaseItemOut(
                id=item.id,
                ingredient_id=item.ingredient_id,
                ingredient_name=item.ingredient.name,
                quantity=item.quantity,
                unit=item.unit,
                unit_price=item.unit_price,
                amount=item.amount,
            )
            for item in purchase.items
        ],
    )


def get_purchase_or_404(db: Session, purchase_id: int):
    purchase = (
        db.query(Purchase)
        .options(joinedload(Purchase.items).joinedload(PurchaseItem.ingredient))
        .filter(Purchase.id == purchase_id)
        .first()
    )
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase not found")
    return purchase


def apply_purchase_items(db: Session, purchase: Purchase, payload: PurchaseCreate):
    purchase.purchase_date = payload.purchase_date
    purchase.items.clear()
    total = Decimal("0")
    for incoming in payload.items:
        ingredient = db.get(Ingredient, incoming.ingredient_id)
        if not ingredient:
            raise HTTPException(status_code=404, detail=f"Ingredient {incoming.ingredient_id} not found")
        amount = money(incoming.quantity * incoming.unit_price)
        purchase.items.append(
            PurchaseItem(
                ingredient_id=ingredient.id,
                quantity=incoming.quantity,
                unit=ingredient.unit,
                unit_price=incoming.unit_price,
                amount=amount,
            )
        )
        total += amount
    purchase.total_amount = money(total)
    purchase.updated_at = datetime.now(timezone.utc)


@app.on_event("startup")
def startup():
    Base.metadata.create_all(bind=engine)
    db = next(get_db())
    try:
        if db.query(Ingredient).count() == 0:
            for name, unit in DEFAULT_INGREDIENTS:
                db.add(Ingredient(name=name, unit=unit))
            db.commit()
    finally:
        db.close()


@app.get("/health")
def health():
    return {"status": "ok", "date": date.today().isoformat()}


@app.get("/ingredients", response_model=list[IngredientOut])
def list_ingredients(db: Session = Depends(get_db)):
    return db.query(Ingredient).order_by(Ingredient.name.asc()).all()


@app.post("/ingredients", response_model=IngredientOut)
def create_ingredient(payload: IngredientCreate, db: Session = Depends(get_db)):
    exists = db.query(Ingredient).filter(func.lower(Ingredient.name) == payload.name.lower()).first()
    if exists:
        raise HTTPException(status_code=409, detail="Ingredient already exists")
    ingredient = Ingredient(name=payload.name.strip(), unit=payload.unit.strip())
    db.add(ingredient)
    db.commit()
    db.refresh(ingredient)
    return ingredient


@app.put("/ingredients/{ingredient_id}", response_model=IngredientOut)
def update_ingredient(ingredient_id: int, payload: IngredientUpdate, db: Session = Depends(get_db)):
    ingredient = db.get(Ingredient, ingredient_id)
    if not ingredient:
        raise HTTPException(status_code=404, detail="Ingredient not found")
    ingredient.name = payload.name.strip()
    ingredient.unit = payload.unit.strip()
    db.commit()
    db.refresh(ingredient)
    return ingredient


@app.delete("/ingredients/{ingredient_id}")
def delete_ingredient(ingredient_id: int, db: Session = Depends(get_db)):
    ingredient = db.get(Ingredient, ingredient_id)
    if not ingredient:
        raise HTTPException(status_code=404, detail="Ingredient not found")
    used = db.query(PurchaseItem).filter(PurchaseItem.ingredient_id == ingredient_id).count()
    if used:
        raise HTTPException(status_code=409, detail="Cannot delete ingredient that has purchase history")
    db.delete(ingredient)
    db.commit()
    return {"deleted": True}


@app.get("/purchases", response_model=list[PurchaseOut])
def list_purchases(
    start_date: date | None = None,
    end_date: date | None = None,
    ingredient: str | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(Purchase).options(joinedload(Purchase.items).joinedload(PurchaseItem.ingredient))
    if start_date:
        query = query.filter(Purchase.purchase_date >= start_date)
    if end_date:
        query = query.filter(Purchase.purchase_date <= end_date)
    if ingredient:
        query = query.join(PurchaseItem).join(Ingredient).filter(Ingredient.name.contains(ingredient))
    purchases = query.order_by(Purchase.purchase_date.desc(), Purchase.created_at.desc()).all()
    return [serialize_purchase(p) for p in purchases]


@app.get("/purchases/{purchase_id}", response_model=PurchaseOut)
def get_purchase(purchase_id: int, db: Session = Depends(get_db)):
    return serialize_purchase(get_purchase_or_404(db, purchase_id))


@app.post("/purchases", response_model=PurchaseOut)
def create_purchase(payload: PurchaseCreate, db: Session = Depends(get_db)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Purchase must include at least one item")
    purchase = Purchase(purchase_date=payload.purchase_date)
    db.add(purchase)
    apply_purchase_items(db, purchase, payload)
    db.commit()
    db.refresh(purchase)
    return serialize_purchase(get_purchase_or_404(db, purchase.id))


@app.put("/purchases/{purchase_id}", response_model=PurchaseOut)
def update_purchase(purchase_id: int, payload: PurchaseCreate, db: Session = Depends(get_db)):
    if not payload.items:
        raise HTTPException(status_code=400, detail="Purchase must include at least one item")
    purchase = get_purchase_or_404(db, purchase_id)
    apply_purchase_items(db, purchase, payload)
    db.commit()
    return serialize_purchase(get_purchase_or_404(db, purchase_id))


@app.delete("/purchases/{purchase_id}")
def delete_purchase(purchase_id: int, db: Session = Depends(get_db)):
    purchase = get_purchase_or_404(db, purchase_id)
    db.delete(purchase)
    db.commit()
    return {"deleted": True}


@app.get("/dashboard", response_model=DashboardOut)
def dashboard(db: Session = Depends(get_db)):
    today = date.today()
    month_start = today.replace(day=1)
    next_month = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)

    today_purchases = db.query(Purchase).filter(Purchase.purchase_date == today).all()
    today_total = sum((money(p.total_amount) for p in today_purchases), Decimal("0"))
    today_items = (
        db.query(PurchaseItem)
        .join(Purchase)
        .join(Ingredient)
        .filter(Purchase.purchase_date == today)
        .all()
    )
    today_ingredients = sorted({item.ingredient.name for item in today_items})

    month_purchases = (
        db.query(Purchase)
        .filter(and_(Purchase.purchase_date >= month_start, Purchase.purchase_date < next_month))
        .all()
    )
    month_total = sum((money(p.total_amount) for p in month_purchases), Decimal("0"))
    month_item_count = (
        db.query(PurchaseItem)
        .join(Purchase)
        .filter(and_(Purchase.purchase_date >= month_start, Purchase.purchase_date < next_month))
        .count()
    )

    return DashboardOut(
        today_total=money(today_total),
        today_item_count=len(today_items),
        today_ingredients=today_ingredients,
        month_total=money(month_total),
        ingredient_count=db.query(Ingredient).count(),
        month_recorded_days=len({p.purchase_date for p in month_purchases}),
        month_item_count=month_item_count,
    )


@app.get("/reports", response_model=ReportOut)
def reports(
    scope: str = Query("month", pattern="^(day|week|month|year)$"),
    start_date: date | None = None,
    end_date: date | None = None,
    db: Session = Depends(get_db),
):
    today = date.today()
    if not start_date or not end_date:
        if scope == "day":
            start_date = end_date = today
        elif scope == "week":
            start_date = today - timedelta(days=today.weekday())
            end_date = start_date + timedelta(days=6)
        elif scope == "month":
            start_date = today.replace(day=1)
            end_date = ((start_date.replace(day=28) + timedelta(days=4)).replace(day=1)) - timedelta(days=1)
        else:
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)

    purchases = (
        db.query(Purchase)
        .filter(and_(Purchase.purchase_date >= start_date, Purchase.purchase_date <= end_date))
        .order_by(Purchase.purchase_date.asc())
        .all()
    )
    totals_by_label = {}
    counts_by_label = {}
    for purchase in purchases:
        label = purchase.purchase_date.strftime("%Y-%m-%d" if scope in {"day", "week", "month"} else "%Y-%m")
        totals_by_label[label] = totals_by_label.get(label, Decimal("0")) + money(purchase.total_amount)
        counts_by_label[label] = counts_by_label.get(label, 0) + 1

    item_count = (
        db.query(PurchaseItem)
        .join(Purchase)
        .filter(and_(Purchase.purchase_date >= start_date, Purchase.purchase_date <= end_date))
        .count()
    )

    top_rows = (
        db.query(
            Ingredient.name,
            func.count(PurchaseItem.id),
            func.coalesce(func.sum(PurchaseItem.quantity), 0),
            func.coalesce(func.sum(PurchaseItem.amount), 0),
        )
        .join(PurchaseItem, PurchaseItem.ingredient_id == Ingredient.id)
        .join(Purchase, Purchase.id == PurchaseItem.purchase_id)
        .filter(and_(Purchase.purchase_date >= start_date, Purchase.purchase_date <= end_date))
        .group_by(Ingredient.id)
        .order_by(func.count(PurchaseItem.id).desc())
        .limit(10)
        .all()
    )

    return ReportOut(
        scope=scope,
        total_amount=money(sum((money(p.total_amount) for p in purchases), Decimal("0"))),
        item_count=item_count,
        recorded_days=len({p.purchase_date for p in purchases}),
        series=[
            ReportPoint(label=label, total=money(total), count=counts_by_label[label])
            for label, total in sorted(totals_by_label.items())
        ],
        top_ingredients=[
            TopIngredient(
                ingredient_name=row[0],
                purchase_count=row[1],
                total_quantity=row[2],
                total_amount=money(row[3]),
            )
            for row in top_rows
        ],
    )


@app.get("/export/excel")
def export_excel(
    start_date: date | None = None,
    end_date: date | None = None,
    db: Session = Depends(get_db),
):
    query = db.query(PurchaseItem).join(Purchase).join(Ingredient)
    if start_date:
        query = query.filter(Purchase.purchase_date >= start_date)
    if end_date:
        query = query.filter(Purchase.purchase_date <= end_date)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventory Purchases"
    sheet.append(["วันที่", "เวลา", "วัตถุดิบ", "จำนวน", "หน่วย", "ราคา", "ยอดรวม"])
    for item in query.order_by(Purchase.purchase_date.asc()).all():
        sheet.append(
            [
                item.purchase.purchase_date.isoformat(),
                item.purchase.created_at.strftime("%H:%M"),
                item.ingredient.name,
                float(item.quantity),
                item.unit,
                float(item.unit_price),
                float(item.amount),
            ]
        )
    export_dir = Path(__file__).resolve().parent / "exports"
    export_dir.mkdir(exist_ok=True)
    file_path = export_dir / f"inventory-export-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    workbook.save(file_path)
    return FileResponse(file_path, filename=file_path.name)


@app.post("/backup")
def backup_database(db: Session = Depends(get_db)):
    backup_dir = Path(__file__).resolve().parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    target = backup_dir / f"inventory-backup-{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"

    workbook = Workbook()
    ingredients_sheet = workbook.active
    ingredients_sheet.title = "ingredients"
    ingredients_sheet.append(["id", "name", "unit", "created_at"])
    for ingredient in db.query(Ingredient).order_by(Ingredient.id.asc()).all():
        ingredients_sheet.append([ingredient.id, ingredient.name, ingredient.unit, ingredient.created_at.isoformat()])

    purchases_sheet = workbook.create_sheet("purchases")
    purchases_sheet.append(["id", "purchase_date", "total_amount", "created_at", "updated_at"])
    for purchase in db.query(Purchase).order_by(Purchase.id.asc()).all():
        purchases_sheet.append(
            [
                purchase.id,
                purchase.purchase_date.isoformat(),
                float(purchase.total_amount),
                purchase.created_at.isoformat(),
                purchase.updated_at.isoformat(),
            ]
        )

    items_sheet = workbook.create_sheet("purchase_items")
    items_sheet.append(["id", "purchase_id", "ingredient_id", "quantity", "unit", "unit_price", "amount"])
    for item in db.query(PurchaseItem).order_by(PurchaseItem.id.asc()).all():
        items_sheet.append(
            [
                item.id,
                item.purchase_id,
                item.ingredient_id,
                float(item.quantity),
                item.unit,
                float(item.unit_price),
                float(item.amount),
            ]
        )

    workbook.save(target)
    return {"backup_file": str(target)}
